from openpyxl import *
import os
import pyautogui as pag
thisPC = os.environ["COMPUTERNAME"]
#thisPC = "ATS1-4"

wb = load_workbook(filename="ATSbotDB.xlsx", read_only=False)
ws = wb.active

model = "EUG150S350"

def printValues():
    print("Host -> ATS Type")
    for row in range(2, ws.max_row+1):
        currentRow = "A"+str(row)
        atsModel = "B"+str(row)
        print("{} -> {}".format(ws[currentRow].value,int(ws[atsModel].value)))

def findHostCell(hostName):
    print("Looking for host. Max row -> "+str(ws.max_row))
    for row in range(1,ws.max_row+1):
        #look for host
        hostRow = "A"+str(row)
        if ws[hostRow].value == hostName:
            print(ws[hostRow].value +  " POSITION IS-> "+hostRow+ " is equal -> "+str((hostName == ws[hostRow].value)))
            return hostRow
    return None

def writeModelInDB(modelName, hostName):
  if findHostCell(hostName) is not None:
    currentRow ="C" + findHostCell(hostName)[-1]
    ws[currentRow] = modelName
    wb.save("ATSbotDB.xlsx")
    print("File saved")
    return True
  else:
    return False

def writeModelInDB(modelName):
  if findHostCell(thisPC) is not None:
    currentRow ="C" + findHostCell(thisPC)[1:]
    ws[currentRow] = modelName
    wb.save("ATSbotDB.xlsx")
    print("Added "+modelName)
    print("File saved")
    return True
  else:
    print('Error')
    return False


def getModelFromDB():
    if findHostCell(thisPC) is not None:
        modelRow = "C" + findHostCell(thisPC)[1:]
        print("Model Row is "+modelRow)
        return ws[modelRow].value
    else:
        return None


def getAtsTypeFromDB(hostName):
    if findHostCell(hostName) is not None:
        modelRow = "B" + findHostCell(hostName)[-1]
        return ws[modelRow].value
    else:
        return None

def getAtsTypeFromDB():
    if findHostCell(thisPC) is not None:
        modelRow = "B" + findHostCell(thisPC)[1:]
        return ws[modelRow].value
    else:
        return None







